'''
This py file is aimed to convert the taxi receipt to excel file,
the crucial OCR technique is realized by many company like Baidu,
as i just use the technique for myself,and only have about 20 receipts each month at most,
so here i just call the baidu OCR-API(https://cloud.baidu.com/doc/OCR/OCR-API.html) to achieve my convert goal.
if you use the Baidu OCR-API for business, you must pay for it to get enough number of calls.
'''

# coding=utf-8
__author__ = 'lulu wang'
__author_email__ = 'gzgzemail@sina.com'
__version__ = '0.1.1dev'

import json
import base64
import requests
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment

#to call the OCR-API, one must sign in BaiduYun,get related access_token and session_key
#you can find the tutorial in http://ai.baidu.com/forum/topic/show/867951
#token_key_dir = "your json file saved token and key"
def get_token_key(token_key_dir):
    parameter = None
    with open(token_key_dir,"rb") as f:
        parameter = json.loads(f.read())
    return parameter

#images_dir is the dir where taxi-receipts located
#image type must in ("PNG、JPG、JPEG、BMP" )
#return each image's path
def get_images(images_dir):
    images=[] 
    for root, dirs, files in os.walk(images_dir):
        for file in files:
            if (os.path.splitext(file)[1]).lower() in ['.jpeg','.png','.jpg','.bmp']:
                images.append(os.path.join(root, file))
    return images

#the receipt image must base64 encoded
#images_dir is the dir where the taxi-receipts located
#return the base64 encoded image
def image_encode(images_dir):
    images_name = get_images(images_dir)
    base64_images = []
    for name in images_name:
        with open(name,"rb") as f:
            image_data = f.read()
            base64_images.append(base64.b64encode(image_data))  
    return base64_images

#make the post requests to call the OCR-API and get the OCR result
def get_ocr_result(parameter,base64_images, name):
    url = "https://aip.baidubce.com/rest/2.0/ocr/v1/taxi_receipt?access_token=%s"%(parameter["access_token"])
    headers = {'content-type': "application/x-www-form-urlencoded"}
    response_list = []
    for image in range(0,len(base64_images)):
        image_data = base64_images[image].decode("utf-8")
        body = { "image_type":"BASE64","image":image_data, "group_id": "gropu001", "user_id": "0001"}
        response = requests.post(url, data = body, headers = headers)
        result = json.loads(response.text)
        if response.status_code==200:
            time = result['words_result']['Time'].split('-')[0]
            date = result['words_result']['Date']
            fare = round(float(result['words_result']['Fare'][1:6]))+int(result['words_result']['FuelOilSurcharge'][1])
            response_list.append([date, name,None, None, None, fare])
    return response_list

#save the result to xlsx file
#the format of my xlsx is as excel.png shows
def save2xlsx(response_list, xlsx_file, start_row, start_column, end_row, end_column):
    al = Alignment(horizontal="center", vertical="center")
    wb = load_workbook(xlsx_file)
    ws = wb['sheet1']
    cells = ws["%s%d"%(start_column,start_row):"%s%d"%(end_column,end_row)]
    index = 0
    for c1,c2,c3,c4,c5,c6 in cells:
        c1.value = response_list[index][0]
        c2.value = response_list[index][1]
        c3.value = response_list[index][2]
        c4.value = response_list[index][3]
        c5.value = response_list[index][4]
        c6.value = response_list[index][5]
        c1.alignment = al
        c2.alignment = al
        c6.alignment = al
        index += 1
    wb.save(xlsx_file)
    print('ok!')
	
if __name__=="__main__":
    token_key_dir = "F:/tmp/params.txt"
    parameter = get_token_key(token_key_dir)
    images_dir = "F:/OCR/images"
    image_data = image_encode(images_dir)
    name = "张三"
    response_list = get_ocr_result(parameter,image_data,name)
    xlsx_file = 'F:/tmp/test.xlsx'
    start_row = 4
    start_column = 'B'
    end_row = len(response_list)+start_row-1
    end_column = 'G'
    save2xlsx(response_list,xlsx_file,start_row, start_column, end_row, end_column)
